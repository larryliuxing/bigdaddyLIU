#!/usr/bin/env python3
"""Generate office reports from GitHub PR metadata."""

from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "pr-source.json"
CST = timezone(timedelta(hours=8))

# 已知 PR 按业务线人工归类，避免「拍卖」「后台」等词交叉误伤。
PR_MODULE = {
    1: "登录 / 基础平台",
    2: "登录 / 基础平台",
    3: "排行榜 / OCR",
    4: "成员 / 权限",
    5: "拍卖竞价",
    6: "拍卖竞价",
    7: "拍卖竞价",
    8: "拍卖竞价",
    9: "拍卖竞价",
    10: "拍卖竞价",
    11: "拍卖竞价",
    12: "BOSS 计时",
    13: "拍卖竞价",
    14: "首页 / 品牌",
    15: "成员 / 权限",
    16: "BOSS 计时",
    17: "BOSS 计时",
    18: "BOSS 计时",
    19: "上线 / 运维",
    20: "排行榜 / OCR",
    21: "排行榜 / OCR",
    22: "排行榜 / OCR",
    23: "排行榜 / OCR",
    24: "拍卖竞价",
    25: "拍卖竞价",
    26: "拍卖竞价",
    27: "排行榜 / OCR",
    28: "排行榜 / OCR",
    29: "首页 / 品牌",
    30: "排行榜 / OCR",
    31: "首页 / 品牌",
    32: "排行榜 / OCR",
    33: "排行榜 / OCR",
    34: "排行榜 / OCR",
    35: "排行榜 / OCR",
    36: "BOSS 计时",
    37: "BOSS 计时",
    38: "拍卖竞价",
    39: "首页 / 品牌",
    40: "拍卖竞价",
    41: "拍卖竞价",
    42: "排行榜 / OCR",
    43: "排行榜 / OCR",
    44: "排行榜 / OCR",
}

MODULE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("BOSS 计时", ("boss", "倒计时")),
    ("排行榜 / OCR", ("排行榜", "战力", "点名", "ocr", "leaderboard", "hud")),
    ("拍卖竞价", ("拍卖", "拍品", "竞拍", "分红", "税率", "fanfare", "auction", "粉色")),
    ("首页 / 品牌", ("首页", "公告", "更新日志", "天堂2", "brand")),
    ("登录 / 基础平台", ("登录", "cookie", "login")),
    ("成员 / 权限", ("成员", "清退", "权限")),
    ("上线 / 运维", ("上线", "go-live", "cleanup")),
]


def classify(pr: dict) -> str:
    number = pr.get("number")
    if number in PR_MODULE:
        return PR_MODULE[number]
    t = pr["title"].lower()
    for module, keys in MODULE_RULES:
        if any(k.lower() in t for k in keys):
            return module
    return "其他"


def type_of(title: str) -> str:
    t = title.lower()
    if t.startswith("fix") or title.startswith("修复") or "fix:" in t:
        return "缺陷修复"
    if t.startswith("perf") or "加速" in title or "加快" in title:
        return "性能优化"
    if t.startswith("chore") or "清理" in title:
        return "运维清理"
    if t.startswith("feat") or title.startswith("feat"):
        return "功能开发"
    if any(k in title for k in ("feat", "可", "支持", "改为", "重构", "增加", "加大")):
        return "功能开发"
    return "功能开发"


def parse_dt(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(CST)


def week_label(dt: datetime) -> str:
    monday = dt.date() - timedelta(days=dt.weekday())
    sunday = monday + timedelta(days=6)
    return f"{monday.isoformat()} ~ {sunday.isoformat()}"


def load_prs() -> list[dict]:
    rows = json.loads(SOURCE.read_text(encoding="utf-8"))
    out = []
    for pr in rows:
        created = parse_dt(pr["createdAt"])
        updated = parse_dt(pr["updatedAt"])
        title = pr["title"]
        out.append(
            {
                "编号": pr["number"],
                "标题": title,
                "模块": classify(pr),
                "类型": type_of(title),
                "状态": "草稿" if pr["isDraft"] else ("已合并" if pr["mergedAt"] else "待审"),
                "草稿": "是" if pr["isDraft"] else "否",
                "创建日期": created.date().isoformat(),
                "创建时间": created.strftime("%Y-%m-%d %H:%M"),
                "更新日期": updated.date().isoformat(),
                "更新时间": updated.strftime("%Y-%m-%d %H:%M"),
                "周次": week_label(created),
                "源分支": pr["headRefName"],
                "目标分支": pr["baseRefName"],
                "链接": pr["url"],
            }
        )
    out.sort(key=lambda r: r["编号"])
    return out


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, name="微软雅黑", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[row].height = 22


def autosize(ws, min_w=10, max_w=42) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            if cell.value is None:
                continue
            width = max(width, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


THIN = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)


def write_table(ws, headers, rows, table_name: str) -> None:
    ws.append(headers)
    header_row = ws.max_row
    style_header(ws, header_row)
    zebra = PatternFill("solid", fgColor="F6F8FB")
    for offset, row in enumerate(rows):
        ws.append(row)
        i = header_row + 1 + offset
        for cell in ws[i]:
            cell.border = THIN
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if offset % 2 == 0:
                cell.fill = zebra
        ws.row_dimensions[i].height = 20
    end = header_row + len(rows)
    ref = f"A{header_row}:{get_column_letter(len(headers))}{end}"
    ws.add_table(
        Table(
            displayName=table_name,
            ref=ref,
            tableStyleInfo=TableStyleInfo(name="TableStyleMedium2", showRowStripes=True),
        )
    )
    autosize(ws)
    ws.auto_filter.ref = ref
    if header_row == 1:
        ws.freeze_panes = "A2"


def build_excel(prs: list[dict]) -> None:
    wb = Workbook()

    ws_cover = wb.active
    ws_cover.title = "封面"
    ws_cover["A1"] = "天堂2盟约 · 公会助手"
    ws_cover["A1"].font = Font(name="微软雅黑", size=22, bold=True, color="1F4E79")
    ws_cover["A2"] = "研发进度办公报表"
    ws_cover["A2"].font = Font(name="微软雅黑", size=16, color="2E75B6")
    meta = [
        ("报告日期", datetime.now(CST).strftime("%Y-%m-%d")),
        ("统计区间", f"{prs[0]['创建日期']} 至 {prs[-1]['创建日期']}"),
        ("数据来源", "GitHub Pull Requests（bigdaddyLIU）"),
        ("PR 总数", len(prs)),
        ("已合并到 main", sum(1 for p in prs if p["目标分支"] == "main" and p["状态"] == "已合并")),
        ("当前状态", "全部 44 个 PR 仍为 Open，且多为串联叠加分支"),
        ("说明", "本表用于周会/复盘。main 仍只有 README，功能代码在功能分支链上。"),
    ]
    ws_cover["A4"] = "字段"
    ws_cover["B4"] = "内容"
    style_header(ws_cover, 4)
    for i, (k, v) in enumerate(meta, start=5):
        ws_cover[f"A{i}"] = k
        ws_cover[f"B{i}"] = v
        ws_cover[f"A{i}"].font = Font(name="微软雅黑", bold=True)
        ws_cover[f"B{i}"].font = Font(name="微软雅黑")
        ws_cover[f"B{i}"].alignment = Alignment(wrap_text=True)
    ws_cover.column_dimensions["A"].width = 22
    ws_cover.column_dimensions["B"].width = 78
    ws_cover.row_dimensions[11].height = 36

    headers = [
        "编号",
        "标题",
        "模块",
        "类型",
        "状态",
        "草稿",
        "创建日期",
        "更新时间",
        "周次",
        "源分支",
        "目标分支",
        "链接",
    ]
    ws_inv = wb.create_sheet("PR明细")
    write_table(
        ws_inv,
        headers,
        [[p[h] for h in headers] for p in prs],
        "PRInventory",
    )

    module_counts = Counter(p["模块"] for p in prs)
    type_counts = Counter(p["类型"] for p in prs)
    week_counts = Counter(p["周次"] for p in prs)
    status_counts = Counter(p["状态"] for p in prs)

    ws_mod = wb.create_sheet("模块汇总")
    write_table(
        ws_mod,
        ["模块", "PR数量", "占比"],
        [
            [k, v, round(v / len(prs), 3)]
            for k, v in sorted(module_counts.items(), key=lambda x: -x[1])
        ],
        "ModuleSummary",
    )
    pie = PieChart()
    pie.title = "PR 模块分布"
    labels = Reference(ws_mod, min_col=1, min_row=2, max_row=1 + len(module_counts))
    data = Reference(ws_mod, min_col=2, min_row=1, max_row=1 + len(module_counts))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.width = 14
    pie.height = 8
    ws_mod.add_chart(pie, "E2")

    ws_week = wb.create_sheet("周度产出")
    week_rows = sorted(week_counts.items())
    write_table(ws_week, ["周次", "PR数量"], [[k, v] for k, v in week_rows], "WeeklyOutput")
    bar = BarChart()
    bar.title = "每周新建 PR 数量"
    bar.y_axis.title = "PR 数"
    bar.add_data(Reference(ws_week, min_col=2, min_row=1, max_row=1 + len(week_rows)), titles_from_data=True)
    bar.set_categories(Reference(ws_week, min_col=1, min_row=2, max_row=1 + len(week_rows)))
    bar.shape = 4
    bar.width = 18
    bar.height = 9
    ws_week.add_chart(bar, "E2")

    ws_type = wb.create_sheet("类型与状态")
    write_table(
        ws_type,
        ["分类", "项目", "数量"],
        [["类型", k, v] for k, v in type_counts.most_common()]
        + [["状态", k, v] for k, v in status_counts.most_common()],
        "TypeStatus",
    )

    chain = wb.create_sheet("分支链路")
    chain.append(["风险说明", "当前 44 个 PR 基本首尾相接（后一个 PR 的目标分支 = 前一个 PR 的源分支）。main 未被合入，功能实际堆积在最长功能分支上。合并或回滚成本会随链路变长而上升。"])
    chain["A1"].font = Font(name="微软雅黑", color="9C2B2B", bold=True)
    chain.merge_cells("A1:F1")
    chain.row_dimensions[1].height = 40
    chain.append([])
    write_table(
        chain,
        ["编号", "源分支", "目标分支", "标题"],
        [[p["编号"], p["源分支"], p["目标分支"], p["标题"]] for p in prs],
        "BranchChain",
    )

    wb.save(ROOT / "开发进度分析.xlsx")


def build_csv(prs: list[dict]) -> None:
    headers = [
        "编号",
        "标题",
        "模块",
        "类型",
        "状态",
        "草稿",
        "创建日期",
        "更新时间",
        "周次",
        "源分支",
        "目标分支",
        "链接",
    ]
    with (ROOT / "pr-inventory.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()
        for p in prs:
            writer.writerow({k: p[k] for k in headers})


def md_escape(text: str) -> str:
    return text.replace("|", "\\|")


def build_weekly_md(prs: list[dict]) -> None:
    module_counts = Counter(p["模块"] for p in prs)
    week_counts = Counter(p["周次"] for p in prs)
    drafts = sum(1 for p in prs if p["草稿"] == "是")
    latest = [p for p in prs if p["编号"] >= 36]
    today = datetime.now(CST).strftime("%Y年%m月%d日")

    lines = [
        "# 天堂2盟约公会助手 · 研发进度周报",
        "",
        f"- **报告日期**：{today}",
        f"- **统计区间**：{prs[0]['创建日期']} ~ {prs[-1]['创建日期']}",
        "- **数据来源**：GitHub `larryliuxing/bigdaddyLIU` 全部 Pull Request",
        "- **报告用途**：周会同步、功能盘点、合并风险评估",
        "",
        "## 一、结论先行",
        "",
        f"1. 过去三周共产生 **{len(prs)}** 个 PR，覆盖拍卖、排行榜识别、BOSS 计时、成员后台、首页公告五条业务线。",
        "2. **main 仍只有 README**，功能代码全部留在互相叠加的功能分支上；目前 **0 个 PR 已合并**。",
        f"3. 草稿 PR **{drafts}** 个、待审非草稿 **{len(prs) - drafts}** 个。链路呈「火车式」串联，越往后合并成本越高。",
        "4. 近期（#36–#44）集中在：BOSS 点击即重开倒计时、后台截图批量校时、粉色拍品投票/掷点、拍卖列表加速，以及若干「restore」类修复。",
        "",
        "## 二、关键指标",
        "",
        "| 指标 | 数值 |",
        "| --- | --- |",
        f"| PR 总数 | {len(prs)} |",
        f"| 已合并 | 0 |",
        f"| 草稿 | {drafts} |",
        f"| 待审（非草稿） | {len(prs) - drafts} |",
        f"| 覆盖模块 | {len(module_counts)} |",
        f"| 最活跃周 | {max(week_counts.items(), key=lambda x: x[1])[0]}（{max(week_counts.values())} 个） |",
        "",
        "## 三、模块分布",
        "",
        "| 模块 | PR 数 | 占比 |",
        "| --- | ---: | ---: |",
    ]
    for name, count in sorted(module_counts.items(), key=lambda x: -x[1]):
        lines.append(f"| {name} | {count} | {count / len(prs):.0%} |")

    lines += [
        "",
        "## 四、周度产出",
        "",
        "| 周次（周一至周日，北京时间） | 新建 PR |",
        "| --- | ---: |",
    ]
    for week, count in sorted(week_counts.items()):
        lines.append(f"| {week} | {count} |")

    lines += [
        "",
        "## 五、本周/近期事项（#36 起）",
        "",
        "| PR | 标题 | 模块 | 状态 |",
        "| ---: | --- | --- | --- |",
    ]
    for p in latest:
        lines.append(
            f"| #{p['编号']} | {md_escape(p['标题'])} | {p['模块']} | {p['状态']} |"
        )

    by_module: dict[str, list[dict]] = defaultdict(list)
    for p in prs:
        by_module[p["模块"]].append(p)

    lines += ["", "## 六、按模块整理的工作清单", ""]
    for module, items in sorted(by_module.items(), key=lambda x: -len(x[1])):
        lines.append(f"### {module}（{len(items)}）")
        lines.append("")
        for p in items:
            lines.append(f"- #{p['编号']} {p['标题']} — {p['创建日期']} · {p['状态']}")
        lines.append("")

    lines += [
        "## 七、风险与建议",
        "",
        "1. **分支火车过长**：#2 以 main 为基底后，后续 PR 大多以前一个功能分支为 base。任何中段冲突都会向后传递。",
        "2. **main 空窗**：线上若要以 main 为准，当前无法发布完整功能；需要一次有计划的合入，而不是继续叠 PR。",
        "3. **重复 restore**：#39、#42、#43、#44 标题均带 restore，说明中间改动曾冲掉过首页公告、排行榜阈值、移除记录和青色 OCR。合入时要做回归清单。",
        "4. **建议下一步（办公侧）**：用本目录的 Excel 做周会材料；用《功能模块说明书》做验收对照；合并前先冻结新需求一天，把链路压到 1–2 条主干。",
        "",
        "## 八、附件",
        "",
        "- [开发进度分析.xlsx](./开发进度分析.xlsx) — 明细、模块饼图、周度柱状图、分支链路",
        "- [pr-inventory.csv](./pr-inventory.csv) — 可导入 WPS / 飞书表格",
        "- [功能模块说明书.md](./功能模块说明书.md) — 面向使用/验收的功能说明",
        "- [项目进度看板.html](./项目进度看板.html) — 浏览器打开即可阅读",
        "",
    ]
    (ROOT / "研发进度周报.md").write_text("\n".join(lines), encoding="utf-8")


def build_spec_md(prs: list[dict]) -> None:
    text = """# 天堂2盟约公会助手 · 功能模块说明书

- **产品名称**：天堂2盟约 · 公会助手
- **文档类型**：功能盘点 / 验收对照
- **整理日期**：%s
- **依据**：仓库功能分支最新提交说明与 44 个 PR 标题（非用户访谈）

## 1. 产品定位

给「天堂2盟约」公会使用的 Web 助手：成员登录后处理日常办公向操作——看公告、交战力截图上榜、打拍卖、看 BOSS 计时；管理员在后台管成员、拍品、分红、排行榜和 BOSS。

## 2. 角色与权限

| 角色 | 能做什么 | 相关工作 |
| --- | --- | --- |
| 盟员 | 登录首页、查看公告、上交战力、参与拍卖、点 BOSS、看分红公示 | #1 #2 #36 |
| 管理员 | 成员清退、拍卖场次/拍品/税率、排行榜合格线与纠错、BOSS 计时与截图校时 | #4 #15 #27 #35 #37 |

账户与拍卖物品管理限制在管理员后台（#4）。登录曾出现 HTTP 下 Cookie 丢失进不了首页（#2）。

## 3. 功能模块

### 3.1 首页与品牌

- 站点品牌为「天堂2盟约」（#14）。
- 首页更新公告/更新日志；曾被后续改动冲掉，#39 将其恢复并避开 HomeHub 重写路径。
- 更新日志位置调整到排行榜下方（#31）。
- 拍卖动态可滚动，结束后结果上移（#29）。

### 3.2 排行榜与战力识别

日常办公场景：成员上传/点选截图，系统识别角色名和战斗力，写入排行榜。

| 能力 | 说明 | PR |
| --- | --- | --- |
| 点选识别 | 先点战力数字，再点名字；战力 4–6 位 | #23 |
| 战力区域 | 以左上角战力为准，取消双校验；识别「战斗力」后数字 | #20 #21 |
| 引擎 | 服务端 PP-OCR | #22 |
| 名字框 | 约为战力框两倍；去掉左右图标干扰；支持「洛丶洛」 | #32 #33 |
| 颜色 | 加强浅蓝；恢复饱和青色（飞飞、抖音绵羊） | #33 #34 #44 |
| 范围与性能 | 加大点名范围后又收回加快，避免又大又卡 | #28 #30 |
| 合格线 | 后台可配置合格战力百分比 | #27 #42 |
| 纠错 | 后台可移除已上榜的错误战力记录 | #35 #43 |

### 3.3 拍卖与分红

这是公会「财务办公」主流程：建场次 → 上拍品 → 同时竞拍 → 结束分红公示。

| 能力 | 说明 | PR |
| --- | --- | --- |
| 场次工作流 | 场次列表 + 拍品管理；时间统一北京时间 | #5 #6 |
| 同时竞拍 | 全部拍品同时开拍，不再一件一件来 | #8 |
| 预约预览 | 未开拍先展示拍品，到点才能出价 | #25 |
| 颜色档位 | 紫色 / 普通粉色 / 特殊粉色 | #40 |
| 加时 | 紫/粉拍品最后一分钟出价加时 | #26 |
| 粉色结算 | 限价竞价后匿名投票，平票掷点 | #38 |
| 税率 | 按场次设置 0%%–10%% | #9 #24 |
| 分红 | 按拍品明细 + 公示总表 | #9 |
| 价格参考 | 添加拍品时展示同名历史高/低/均价 | #11 |
| 名称 OCR | 拍品图只认顶部装备名，并提升短名/非紫名准确率 | #7 #10 |
| 体验 | 高价出价有 fanfare 与弹幕 | #13 |
| 性能 | 拍卖房间、分红、后台列表加载加速 | #41 |

### 3.4 BOSS 计时

公会日常值班：看倒计时、点名标记、投票、管理员改表。

| 能力 | 说明 | PR |
| --- | --- | --- |
| 设置项 | 属性标签、掉落图、管理员计时按钮 | #12 |
| 规则 | 区分「下次刷新」与「击杀」计时 | #17 |
| 列表 | 按倒计时排序、名称加大、可配置投票人数 | #18 |
| 盟员操作 | 点 BOSS 立刻重开倒计时，并展示上次标记的人 | #36 |
| 后台 OCR | 公会截图批量识别并调整 BOSS 计时 | #37 |
| 性能 | 设置页列表加载加速 | #16 |

### 3.5 成员管理

- 清退成员（软退出、清账户与在线数据，保留历史口径按各 PR 演进）。
- 清空成员/拍卖时保留 admin 与 BOSS（#15）。
- 正式上线清理：删测试数据、清拍卖、改管理员密码（#19）。

## 4. 建议的验收清单（办公可用）

打印或导入飞书后逐条打勾：

- [ ] 盟员能登录并进入首页，公告仍在
- [ ] 战力截图能识别普通名、带「丶」的名、青色 HUD 名
- [ ] 后台能改合格线、能删错误上榜记录
- [ ] 预约场次到点前不能出价，到点后全部拍品可同时拍
- [ ] 紫色加价、普通粉色加价、特殊粉色投票/掷点三条路径都能走完
- [ ] 分红表能按拍品出明细，税率按场次
- [ ] 点 BOSS 立即重开倒计时，管理员截图可批量改表
- [ ] 清退成员不会误删 admin 与 BOSS 配置

## 5. 文档维护

重新导出本目录报表：

```bash
gh pr list --state all --limit 100 --json number,title,state,createdAt,updatedAt,mergedAt,isDraft,headRefName,baseRefName,url > docs/office/pr-source.json
python3 docs/office/generate_reports.py
```
""" % datetime.now(CST).strftime("%Y-%m-%d")
    (ROOT / "功能模块说明书.md").write_text(text, encoding="utf-8")


def build_html(prs: list[dict]) -> None:
    module_counts = Counter(p["模块"] for p in prs)
    week_counts = Counter(p["周次"] for p in prs)
    rows = "".join(
        f"<tr><td>#{p['编号']}</td><td>{p['标题']}</td><td>{p['模块']}</td>"
        f"<td>{p['类型']}</td><td>{p['状态']}</td><td>{p['创建日期']}</td></tr>"
        for p in prs
    )
    mods = "".join(
        f"<div class='kpi'><div class='n'>{v}</div><div class='l'>{k}</div></div>"
        for k, v in sorted(module_counts.items(), key=lambda x: -x[1])
    )
    weeks = "".join(
        f"<tr><td>{k}</td><td>{v}</td></tr>" for k, v in sorted(week_counts.items())
    )
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <title>天堂2盟约 · 研发进度看板</title>
  <style>
    body {{ font-family: "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif; margin: 0; background: #f4f6f9; color: #1f2a37; }}
    header {{ background: #1f4e79; color: #fff; padding: 28px 40px; }}
    header p {{ opacity: .85; margin: 8px 0 0; }}
    main {{ padding: 24px 40px 48px; }}
    .kpis {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0 28px; }}
    .kpi {{ background: #fff; border: 1px solid #e5eaf1; border-radius: 10px; padding: 14px 18px; min-width: 120px; }}
    .kpi .n {{ font-size: 28px; font-weight: 700; color: #1f4e79; }}
    .kpi .l {{ color: #667085; font-size: 13px; }}
    table {{ border-collapse: collapse; width: 100%; background: #fff; }}
    th, td {{ border: 1px solid #e5eaf1; padding: 8px 10px; text-align: left; font-size: 13px; }}
    th {{ background: #1f4e79; color: #fff; }}
    tr:nth-child(even) {{ background: #f8fafc; }}
    h2 {{ margin-top: 32px; }}
    .note {{ background: #fff7ed; border: 1px solid #fdba74; padding: 12px 16px; border-radius: 8px; }}
  </style>
</head>
<body>
  <header>
    <h1>天堂2盟约公会助手 · 研发进度看板</h1>
    <p>生成时间 {datetime.now(CST).strftime("%Y-%m-%d %H:%M")}（北京时间） · 共 {len(prs)} 个 PR · 0 个已合并</p>
  </header>
  <main>
    <div class="note">main 目前只有 README。功能都在互相串联的 PR 分支上。本看板用于周会，不替代代码审查。</div>
    <h2>模块数量</h2>
    <div class="kpis">{mods}</div>
    <h2>周度产出</h2>
    <table><thead><tr><th>周次</th><th>PR 数</th></tr></thead><tbody>{weeks}</tbody></table>
    <h2>PR 明细</h2>
    <table>
      <thead><tr><th>编号</th><th>标题</th><th>模块</th><th>类型</th><th>状态</th><th>创建日期</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </main>
</body>
</html>
"""
    (ROOT / "项目进度看板.html").write_text(html, encoding="utf-8")


def main() -> None:
    prs = load_prs()
    build_csv(prs)
    build_excel(prs)
    build_weekly_md(prs)
    build_spec_md(prs)
    build_html(prs)
    print(f"generated {len(prs)} PR office reports in {ROOT}")


if __name__ == "__main__":
    main()
